import openpyxl


#  1) 엑셀 만들기
wb = openpyxl.Workbook()

#  2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

#  3) 데이터 추가하기 ( 행은 숫자 , 열은 영문 대문자로 되어 있음 )
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

#  4) 엑셀 저장하기 / r : 모든 문자를 탈출문자 없이 하나의 문자로 취급
#  아니면 모든 역슬래시를 역슬래시*2 혹은 슬래시로 바꿔야 함
wb.save(r'C:\Users\com\Documents\vscode workspace\python\Crawling_Inflearn_NadoCoding\Inflearn\08_네이버_주식_크롤링\참가자_data.xlsx')


