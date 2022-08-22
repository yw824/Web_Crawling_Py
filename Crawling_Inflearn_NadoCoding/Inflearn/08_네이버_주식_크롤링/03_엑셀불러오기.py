import openpyxl


#  1) 엑셀 만들기
wb = openpyxl.Workbook()
fpath = r'C:\Users\com\Documents\vscode workspace\python\Crawling_Inflearn_NadoCoding\Inflearn\08_네이버_주식_크롤링\참가자_data.xlsx'

#  2) 엑셀 워크시트 불러오기
wb = openpyxl.load_workbook(fpath)
ws = wb['오징어게임']


#  3) 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'


#  4) 데이터 저장하기
wb.save(fpath)
